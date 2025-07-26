from flask import Flask, render_template, request, redirect, url_for, flash,jsonify
import openpyxl
import os
from collections import defaultdict
from datetime import datetime


app = Flask(__name__)
app.secret_key = "your_secret_key"  # Needed for flashing messages

# Hardcoded credentials for demonstration purposes
VALID_USERNAME = "Admin"
VALID_PASSWORD = "123"

# Path for the Excel files
PROPERTY_LIST_FILE = 'property_list.xlsx'
TENANT_INFO_FILE = 'tenant_info.xlsx'
RENT_AGREEMENT_INFO_FILE = 'rent_agreement_info.xlsx'


# Ensure the property list file exists (check if the file is missing, create it)
def create_property_list_file():
    if not os.path.exists(PROPERTY_LIST_FILE):
        # Create a new workbook if it doesn't exist
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Property Name', 'Address', 'Owner Name', 'Mobile Number', 'Property Type'])
        wb.save(PROPERTY_LIST_FILE)

# Ensure the tenant info file exists
def create_tenant_info_file():
    if not os.path.exists(TENANT_INFO_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Property Name', 'Tenant Name', 'Tenant Family Member', 'Mobile Number', 'Adhar Number', 'Start Date', 'Rent Decided', 'Deposit Amount'])
        wb.save(TENANT_INFO_FILE)

# Ensure the rent agreement info file exists
def create_rent_agreement_info_file():
    if not os.path.exists(RENT_AGREEMENT_INFO_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Tenant Name', 'Agreement Start Date', 'Agreement End Date', 'Document'])
        wb.save(RENT_AGREEMENT_INFO_FILE)

# Call the functions to ensure the necessary files exist
create_property_list_file()
create_tenant_info_file()
create_rent_agreement_info_file()

# Function to create property-specific file if it doesn't exist
def create_property_excel(property_name):
    property_file = f"{property_name}.xlsx"
    if not os.path.exists(property_file):
        # Create a new workbook with columns for rent data
        wb_property = openpyxl.Workbook()
        ws_property = wb_property.active
        ws_property.append(['rentAmount','rentMonth' ,'rentReceivedDate','rentReceivedDay','paymentMode','pendingAmount' ])
        wb_property.save(property_file)

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if username == VALID_USERNAME and password == VALID_PASSWORD:
            return redirect(url_for("home"))  # Redirect to the home page
        else:
            flash("Invalid username or password. Please try again.", "error")
            return redirect(url_for("login"))

    return render_template("login.html")

@app.route("/home")
def home():
    return render_template("home.html")

@app.route("/add_property", methods=["GET", "POST"])
def add_property():
    if request.method == "POST":
        # Get data from the form
        property_name = request.form['propertyName']
        address = request.form['address']
        owner_name = request.form['ownerName']
        mobile_number = request.form['mobileNumber']
        property_type = request.form['propertyType']

        # Save the property data to property_list.xlsx
        wb = openpyxl.load_workbook(PROPERTY_LIST_FILE)
        ws = wb.active
        ws.append([property_name, address, owner_name, mobile_number, property_type])
        wb.save(PROPERTY_LIST_FILE)

        # Create a new Excel file for the property (property_name.xlsx) for future rent data
        create_property_excel(property_name)

        flash(f"Property '{property_name}' added successfully!", "success")
        return redirect(url_for('home'))  # Redirect to the home page after success

    return render_template("add_property.html")









@app.route("/add_tenant", methods=["GET", "POST"])
def add_tenant():
    if request.method == "POST":
        # Get tenant data from the form
        property_name = request.form['propertyName']
        tenant_name = request.form['tenantName']
        family_member = request.form['familyMember']
        mobile_number = request.form['mobileNumber']
        adhar_number = request.form['adharNumber']
        start_date = request.form['startDate']
        rent_decided = request.form['rentDecided']
        deposit_amount = request.form['depositAmount']

        # Save tenant data to tenant_info.xlsx
        wb = openpyxl.load_workbook(TENANT_INFO_FILE)
        ws = wb.active
        ws.append([property_name, tenant_name, family_member, mobile_number, adhar_number, start_date, rent_decided, deposit_amount])
        wb.save(TENANT_INFO_FILE)

        flash(f"Tenant '{tenant_name}' added successfully!", "success")
        return redirect(url_for('home'))

    return render_template("add_tenant.html")

@app.route("/add_agreement", methods=["GET", "POST"])
def add_agreement():
    if request.method == "POST":
        tenant_name = request.form['tenantAgreementName']
        agreement_start_date = request.form['agreementStartDate']
        agreement_end_date = request.form['agreementEndDate']
        document = request.files['uploadDocument']

        # Save agreement data to rent_agreement_info.xlsx
        wb = openpyxl.load_workbook(RENT_AGREEMENT_INFO_FILE)
        ws = wb.active
        # Save file to the 'uploads' folder
        upload_folder = 'uploads'
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        
        document_path = os.path.join(upload_folder, document.filename)
        document.save(document_path)

        # Append agreement info to the spreadsheet
        ws.append([tenant_name, agreement_start_date, agreement_end_date, document.filename])
        wb.save(RENT_AGREEMENT_INFO_FILE)

        flash(f"Agreement for '{tenant_name}' added successfully!", "success")
        return redirect(url_for('home'))

    return render_template("add_agreement.html")











# Define the path to the property list file
PROPERTY_LIST_FILE = "property_list.xlsx"
TENANT_INFO_FILE = "tenant_info.xlsx"
RENT_AGREEMENT_INFO_FILE = "rent_agreement_info.xlsx"

@app.route("/view_data")
def view_data():
    # Render the template for viewing data
    return render_template("view.html")

# Rent-related functionality
@app.route("/add_rent", methods=["GET", "POST"])
def add_rent():
    if request.method == "POST":
        # Get the rent data from the form
        property_name = request.form['propertyName']
        rent_amount = request.form['rentAmount']
        rent_month = request.form['rentMonth']
        rent_received_date = request.form['rentReceivedDate']
        rent_received_day = request.form['rentReceivedDay']
        payment_mode = request.form['paymentMode']
        pending_amount = request.form['pendingAmount']

        # Add rent data to the respective property file
        rent_file = f"{property_name}.xlsx"
        if not os.path.exists(rent_file):
            flash(f"The property file '{property_name}.xlsx' does not exist.", "error")
            return redirect(url_for('add_rent'))

        wb = openpyxl.load_workbook(rent_file)
        ws = wb.active
        ws.append([rent_amount, rent_month, rent_received_date, rent_received_day, payment_mode, pending_amount])
        wb.save(rent_file)

        flash(f"Rent data for {property_name} added successfully!", "success")
        return redirect(url_for('home'))

    return render_template("add_rent.html")

@app.route('/get_properties')
def get_properties():
    try:
        wb = openpyxl.load_workbook(PROPERTY_LIST_FILE)
        ws = wb.active
        properties = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
        return jsonify({"properties": properties})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch properties: {str(e)}"})

@app.route('/get_property_data/<property_name>')
def get_property_data(property_name):
    try:
        file_path = f"{property_name}.xlsx"
        if not os.path.exists(file_path):
            return jsonify({"error": f"No data found for property: {property_name}"})

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch data for property '{property_name}': {str(e)}"})

@app.route('/update_row/<property_name>/<int:row_index>', methods=['POST'])
def update_row(property_name, row_index):
    try:
        file_path = f"{property_name}.xlsx"
        if not os.path.exists(file_path):
            return jsonify({"error": f"No data found for property: {property_name}"})

        updated_row = request.json['updatedRow']
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for col_index, value in enumerate(updated_row, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)

        wb.save(file_path)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to update row: {str(e)}"})

@app.route('/delete_row/<property_name>/<int:row_index>', methods=['DELETE'])
def delete_row(property_name, row_index):
    try:
        file_path = f"{property_name}.xlsx"
        if not os.path.exists(file_path):
            return jsonify({"error": f"No data found for property: {property_name}"})

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        ws.delete_rows(row_index + 2)
        wb.save(file_path)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to delete row: {str(e)}"})

# New route to fetch data from property_list.xlsx
@app.route('/get_property_list_data')
def get_property_list_data():
    try:
        wb = openpyxl.load_workbook(PROPERTY_LIST_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch property list data: {str(e)}"})
    
    # Update and delete CRUD functionality for property_list.xlsx
@app.route('/update_property_list_row/<int:row_index>', methods=['POST'])
def update_property_list_row(row_index):
    try:
        updated_row = request.json['updatedRow']
        wb = openpyxl.load_workbook(PROPERTY_LIST_FILE)
        ws = wb.active

        for col_index, value in enumerate(updated_row, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)

        wb.save(PROPERTY_LIST_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to update property list row: {str(e)}"})

@app.route('/delete_property_list_row/<int:row_index>', methods=['DELETE'])
def delete_property_list_row(row_index):
    try:
        wb = openpyxl.load_workbook(PROPERTY_LIST_FILE)
        ws = wb.active

        ws.delete_rows(row_index + 2)
        wb.save(PROPERTY_LIST_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to delete property list row: {str(e)}"})

# New route to fetch data from tenant_info.xlsx
@app.route('/get_tenant_info_data')
def get_tenant_info_data():
    try:
        wb = openpyxl.load_workbook(TENANT_INFO_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch tenant info data: {str(e)}"})
    

# Update and delete CRUD functionality for tenant_info.xlsx
@app.route('/update_tenant_info_row/<int:row_index>', methods=['POST'])
def update_tenant_info_row(row_index):
    try:
        updated_row = request.json['updatedRow']
        wb = openpyxl.load_workbook(TENANT_INFO_FILE)
        ws = wb.active

        for col_index, value in enumerate(updated_row, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)

        wb.save(TENANT_INFO_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to update tenant info row: {str(e)}"})

@app.route('/delete_tenant_info_row/<int:row_index>', methods=['DELETE'])
def delete_tenant_info_row(row_index):
    try:
        wb = openpyxl.load_workbook(TENANT_INFO_FILE)
        ws = wb.active

        ws.delete_rows(row_index + 2)
        wb.save(TENANT_INFO_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to delete tenant info row: {str(e)}"})







# New route to fetch data from rent_agreement_info.xlsx
@app.route('/get_rent_agreement_info_data')
def get_rent_agreement_info_data():
    try:
        wb = openpyxl.load_workbook(RENT_AGREEMENT_INFO_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch rent agreement info data: {str(e)}"})
    

# Update and delete CRUD functionality for tenant_info.xlsx
@app.route('/update_rent_agreement_info_row/<int:row_index>', methods=['POST'])
def update_rent_agreement_info_row(row_index):
    try:
        updated_row = request.json['updatedRow']
        wb = openpyxl.load_workbook(RENT_AGREEMENT_INFO_FILE)
        ws = wb.active

        for col_index, value in enumerate(updated_row, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)

        wb.save(RENT_AGREEMENT_INFO_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to update rent_agreement  info row: {str(e)}"})

@app.route('/delete_rent_greement_info_row/<int:row_index>', methods=['DELETE'])
def delete_rent_greement_info_row(row_index):
    try:
        wb = openpyxl.load_workbook(RENT_AGREEMENT_INFO_FILE)
        ws = wb.active

        ws.delete_rows(row_index + 2)
        wb.save(RENT_AGREEMENT_INFO_FILE)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Failed to delete rent agreement info row: {str(e)}"})




if __name__ == "__main__":
    app.run(debug=True)
